"""
EBORN Microsoft xlsx spreadsheets loader
========================================

Support loading data from multiple sheets, combining them for rendering.

**Prerequisites:** requires `openpyxl <https://pypi.org/project/openpyxl/>_` library

Spreadsheets must contain a column or multiple columns with headers starting 
with ``template_name_key`` argument string. Values of template(s) columns either 
names of the template to use for rendering or OS path string to template file 
relative to ``template_dir`` argument supplied to EBORN object on instantiation.

In addition, table must contain column with ``result_name_key`` values, they used
to combine results, i.e. rendering results for identical ``result_name_key`` combined
in a single string, moreover, if ``file`` uses ``result_name_key`` to name the files
for storing results.
    
Spreadsheet might contain multiple tabs with names starting with ``template``, these 
tabs can contain templates to use for rendering data from other tabs. All the templates
loaded line by line, ``template:{{ template_name }}`` lines used to identify end 
of previous and start of next template, where ``template_name`` used for referencing
template in templates columns.
"""

import logging
from openpyxl import load_workbook

log = logging.getLogger(__name__)


def load_template(sheet, templates_dict):
    current_template_name = ""
    current_template_lines = []
    for rowNum in range(1, sheet.max_row + 1):
        rowData = sheet.cell(row=rowNum, column=1).value
        if rowData != None:  # check if cell is empty and skip it if so
            if rowData.upper().startswith("TEMPLATE:"):
                if current_template_lines and current_template_name:
                    templates_dict[current_template_name] = "\n".join(
                        current_template_lines
                    )
                current_template_name = rowData.split(":")[1].strip()
                current_template_lines = []
            else:
                current_template_lines.append(rowData)
    # add last template
    if current_template_lines and current_template_name:
        templates_dict[current_template_name] = "\n".join(current_template_lines)


def load_data_from_sheet(sheet, ret, template_name_key):
    # headers_split_by_endings = {}
    # headers_with_endings = set()

    headers = [
        sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)
        if not sheet.cell(row=1, column=i).value.startswith("#")
    ]
    
    has_templates_column = False
    for header in headers:
        if header.startswith(template_name_key):
            has_templates_column = True
            break
            
    if not has_templates_column:
        log.warning(
            "XLSX loader, no '{}' header on tab '{}', skipping it".format(
                template_name_key, sheet.name
            )
        )
        return        
        
    # templates_columns_names = {
    #     header: header.replace(template_name_key, "")
    #     for header in headers
    #     if header.startswith(template_name_key)
    # }  # dict {header: ending} where header starts with template_name_key
    # 
    # # skip sheets that does not have template column
    # if not templates_columns_names:
    #     log.warning(
    #         "XLSX loader, no '{}' key in table tab '{}'".format(
    #             template_name_key, sheet_name
    #         )
    #     )
    #     return
    #     
    # # find all headers that has template related endings
    # for template_column, ending in templates_columns_names.items():
    #     if not ending.strip():
    #         continue
    #     for header in headers:
    #         if header.endswith(ending):
    #             headers_with_endings.add(header)
    # headers_without_endings = headers_with_endings.symmetric_difference(headers)
    # 
    # # form headers_split_by_endings lists
    # for template_column, ending in templates_columns_names.items():
    #     headers_split_by_endings[ending] = set(headers_without_endings)
    #     if not ending.strip():
    #         continue
    #     for header in headers:
    #         if header.endswith(ending):
    #             headers_split_by_endings[ending].discard(header[: -len(ending)])
    #             headers_split_by_endings[ending].add(header)
    #             
    # # form data
    # for ending, headers_item in headers_split_by_endings.items():
    #     for row in range(2, sheet.max_row + 1):
    #         ret.append(
    #             {
    #                 h[:-len(ending)]
    #                 if h.endswith(ending) and ending
    #                 else h: sheet.cell(row=row, column=headers.index(h) + 1).value
    #                 for h in headers_item
    #             }
    #         )
    
    # form data
    for row in range(2, sheet.max_row + 1):
        ret.append(
            {
                h: sheet.cell(row=row, column=headers.index(h) + 1).value
                for h in headers
            }
        )


def load(data, templates_dict, template_name_key, **kwargs):
    ret = []

    kwargs.setdefault("data_only", True)
    kwargs.setdefault("read_only", True)

    wb = load_workbook(data, **kwargs)

    for sheet_name in wb.sheetnames:
        log.info("XLSX loader, working with '{}' tab".format(sheet_name))

        if sheet_name.startswith("#"):
            continue
        elif "TEMPLATE" in sheet_name.upper():
            load_template(wb[sheet_name], templates_dict)
        else:
            load_data_from_sheet(wb[sheet_name], ret, template_name_key)
    return ret
