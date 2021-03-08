"""
XLSX Spreadsheets loader plugin
*******************************

**Plugin reference name:** ``xlsx``

This plugin supports loading data from multiple sheets, combining them for rendering.

**Prerequisites:**

- Requires `openpyxl <https://pypi.org/project/openpyxl/>`_ >= 3.0.0 library 

**Restrictions and guidelines**

Spreadsheets must contain a column or multiple columns with headers starting
with ``template_name_key`` argument string, default is ``template``. Values of
template(s) columns either names of the template to use for rendering or OS path
string to template file relative to ``template_dir`` argument supplied to TTR
object on instantiation.

In addition, table must contain column with ``result_name_key`` values, default is
``device``, they used to combine results, i.e. rendering results for identical
``result_name_key`` combined in a single string. ``result_name_key`` used further
by returners to return results.

Spreadsheet tabs with names starting with ``#`` are skipped, useful to comment out
tabs that no need to render.

First row in the spreadsheet must contain headers, otherwise spreadsheet not loaded.

.. note:: empty cells loaded with value of ``None``

Sample spreadsheet table that contains details for interfaces configuration:

+--------+-----------+-----+------+----------+------+--------------------------------------+
| device | interface | vid | vrf  | ip       | mask | template                             |
+========+===========+=====+======+==========+======+======================================+
| rt1    | Gi1/1.100 | 100 | MGMT | 10.0.0.1 | 24   | ttr://simple/interface.cisco_ios.txt |
+--------+-----------+-----+------+----------+------+--------------------------------------+
| rt1    | Gi2/3     |     | CUST | 10.3.0.1 | 30   | ttr://simple/interface.cisco_ios.txt |
+--------+-----------+-----+------+----------+------+--------------------------------------+
| sw23   | Vlan21    |     | MGMT | 10.0.0.2 | 24   | ttr://simple/interface.cisco_ios.txt |
+--------+-----------+-----+------+----------+------+--------------------------------------+

where:

- ``device`` column contains ``result_name_key`` values
- ``template`` column contains ``template_name_key`` values
- ``ttr://simple/interface.cisco_ios.txt``  - is a template included in TTR templates collection

Above table loaded into this list of dictionaries::

    [{'device': 'rt1',
      'interface': 'Gi1/1.100',
      'ip': '10.0.0.1',
      'mask': 24,
      'template': 'ttr://simple/interface.cisco_ios.txt',
      'vid': 100,
      'vrf': 'MGMT'},
     {'device': 'rt1',
      'interface': 'Gi2/3',
      'ip': '10.3.0.1',
      'mask': 30,
      'template': 'ttr://simple/interface.cisco_ios.txt',
      'vid': None,
      'vrf': 'CUST'},
     {'device': 'sw23',
      'interface': 'Vlan21',
      'ip': '10.0.0.2',
      'mask': 24,
      'template': 'ttr://simple/interface.cisco_ios.txt',
      'vid': None,
      'vrf': 'MGMT'}]

Combined with ``ttr://simple/interface.cisco_ios.txt`` it will produce these results::

    ttr -d /path_to_table.xlsx/ -p

    # ---------------------------------------------------------------------------
    # rt1 rendering results
    # ---------------------------------------------------------------------------
    interface Gi1/1.100
     encapsulation dot1q 100
     vrf forwarding  MGMT
     ip address 10.0.0.1 24
     exit
    !
    interface Gi2/3
     encapsulation dot1q None
     vrf forwarding  CUST
     ip address 10.3.0.1 30
     exit
    !

    # ---------------------------------------------------------------------------
    # sw23 rendering results
    # ---------------------------------------------------------------------------
    interface Vlan21
     encapsulation dot1q None
     vrf forwarding  MGMT
     ip address 10.0.0.2 24
     exit
    !

Multiple Templates suffix separation
------------------------------------

Using multitemplate processor it is possible to define multiple template columns
within same spreadsheet tab using suffixes. Columns with headers with same suffixes
considered part of same datum and combined together. Headers without suffixes shared
across all datums.

For example, this table uses ``:a`` and ``:b`` suffixes to denote relationship with certain templates:

+----------+-------------+------------+------+----------+-------------+------------+--------------------------------------+---------------------------------------+
| device:a | interface:a | ip:a       | mask | device:b | interface:b | ip:b       | template:a                           | template:b                            |
+==========+=============+============+======+==========+=============+============+======================================+=======================================+
| rt1      | Gi1/1       | 10.0.0.1   | 30   | rt2      | Gi1         | 10.0.0.2   | ttr://simple/interface.cisco_ios.txt | ttr://simple/interface.cisco_nxos.txt |
+----------+-------------+------------+------+----------+-------------+------------+--------------------------------------+---------------------------------------+
| rt3      | Gi2/3       | 10.3.0.1   | 30   | rt4      | Gi3         | 10.3.0.2   | ttr://simple/interface.cisco_ios.txt | ttr://simple/interface.cisco_nxos.txt |
+----------+-------------+------------+------+----------+-------------+------------+--------------------------------------+---------------------------------------+

where:

- ``device`` columns contains ``result_name_key`` values
- ``template`` columns contains ``template_name_key`` values
- ``ttr://simple/interface.cisco_ios.txt``  - is a template included in TTR templates collection
- ``ttr://simple/interface.cisco_nxos.txt``  - is a template included in TTR templates collection

Above table data, after passing through ``multitemplate`` processor loaded into this list of dictionaries::

    import pprint
    from ttr import ttr

    gen = ttr("./path/to/table.xlsx", processors=["multitemplate"])

    pprint.pprint(gen.data_loaded)

    # prints:
    # [{'device': 'rt1',
    #   'interface': 'Gi1/1',
    #   'ip': '10.0.0.1',
    #   'mask': 30,
    #   'template': 'ttr://simple/interface.cisco_ios.txt'},
    #  {'device': 'rt2',
    #   'interface': 'Gi1',
    #   'ip': '10.0.0.2',
    #   'mask': 30,
    #   'template': 'ttr://simple/interface.cisco_nxos.txt'},
    #  {'device': 'rt3',
    #   'interface': 'Gi2/3',
    #   'ip': '10.3.0.1',
    #   'mask': 30,
    #   'template': 'ttr://simple/interface.cisco_ios.txt'},
    #  {'device': 'rt4',
    #   'interface': 'Gi3',
    #   'ip': '10.3.0.2',
    #   'mask': 30,
    #   'template': 'ttr://simple/interface.cisco_nxos.txt'}]

That technique allows to simplify definition of "paired" configurations, e.g. device A
and device B configs or forward and rollback configurations etc.
"""

import logging
import traceback

from ..templates import templates_loaders_plugins

log = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    log.error("Failed to import openpyxl module")


def load_data_from_sheet(sheet, ret, template_name_key):
    try:
        headers = []
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=i).value == None:
                headers.append(None)
            else:
                headers.append(sheet.cell(row=1, column=i).value)
        # strip spaces if any
        headers = [i.strip() if isinstance(i, str) else i for i in headers]
        # check headers
        if not any(headers):
            log.warning(
                "XLSX loader, sheet '{}' first row is empty, no headers, skipping it.".format(
                    sheet.title
                )
            )
            return
    except:
        log.error(
            "XLSX loader, sheet '{}', failed to load headers, skipping it, error: {}".format(
                sheet.title, traceback.format_exc()
            )
        )
        return

    # check if headers have template column(s)
    has_templates_column = False
    for header in headers:
        if isinstance(header, str) and header.startswith(template_name_key):
            has_templates_column = True
            break


    if not has_templates_column:
        log.warning(
            "XLSX loader, no '{}' header on tab '{}', skipping it".format(
                template_name_key, sheet.title
            )
        )
        return

    # form data
    log.debug("XLSX loader, loading data - tab: '{}', headers: '{}'".format(sheet.title, headers))
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # from data item
        ret.append(
            dict(zip(headers, row))
        )


def load(data, templates_dict, template_name_key, **kwargs):
    """
    Function to load XLSX spreadsheet. Takes OS path to ``.xlsx`` file
    and returns list of dictionaries, where keys equal to headers
    of spreadsheets' tabs.

    :param data: string, OS path to ``.xlsx`` file
    :param templates_dict: dictionary to load templates from spreadsheet
    :param template_name_key: string, templates column header prefix
    """
    ret = []

    kwargs.setdefault("data_only", True)
    kwargs.setdefault("read_only", True)

    wb = load_workbook(data, **kwargs)

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("#"):
            log.debug("XLSX loader, skipping tab - '{}'".format(sheet_name))
            continue
        elif "TEMPLATE" in sheet_name.upper():
            templates_loaders_plugins["xlsx"](templates_dict, sheet=wb[sheet_name])
        else:
            load_data_from_sheet(wb[sheet_name], ret, template_name_key)
    return ret
