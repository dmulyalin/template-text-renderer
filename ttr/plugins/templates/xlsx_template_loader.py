"""
XLSX Template Loader
********************

**Plugin reference name:** ``xlsx``

Spreadsheet might contain multiple tabs with names starting with ``template``, these
tabs can contain templates to use for rendering data from other tabs. All the templates
loaded line by line, ``template:{{ template_name }}`` lines used to identify end
of previous and start of next template, where ``template_name`` used for referencing
template.

Sample table that contains rendering templates, no headers required:

+---------------------------------+
|                                 |                    
+=================================+
| template:interface              |
+---------------------------------+
| interface {{ interface }}       |
+---------------------------------+
|  description {{ descriptiopn }} |
+---------------------------------+
| template:logging                |
+---------------------------------+
| logging host {{ log_server }}   |
+---------------------------------+

Above templates loaded in a dictionary::

    { 
        "interface": 'interface {{ interface }}\\n'
                     '  description {{ descriptiopn }}',
        "logging": 'logging host {{ log_server }}'
    }
    
In this case templates referenced in data using ``interface`` and ``logging`` template names
"""

import os
import logging

log = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    log.error("Failed to import openpyxl module")


def load(templates_dict, templates=None, sheet=None, **kwargs):
    """
    Function to load **all** templates from ``xlsx`` spreadsheets.
    
    :param templates: OS path to ``.xlsx`` file
    :param sheet: openpyxl ``sheet`` object
    :param templates_dict: (dict) dictionary of ``{template name: template content}`` to
        load templates in
    :return: ``True`` on success and ``False`` on failure to load template
    """
    if not templates and not sheet:
        log.error("TTR:xlsx_template_loader, {} - invalid file, {} - invalid sheet, nothing to load.".format(templates, sheet))
        return False
        
    if templates:    
        wb = load_workbook(templates, data_only=True, read_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("#"):
                log.debug("TTR:xlsx_template_loader, skipping tab - '{}'".format(sheet_name))
                continue
            elif "TEMPLATE" in sheet_name.upper():
                load_templates_from_sheet(wb[sheet_name], templates_dict)
    if sheet:
        load_templates_from_sheet(sheet, templates_dict)

    return True
    

def load_templates_from_sheet(sheet, templates_dict):
    log.debug("TTR:xlsx_template_loader, loading templates tab - '{}'".format(sheet.title))
    current_template_name = ""
    current_template_lines = []
    for item in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        rowData = item[0]
        if rowData != None:  # check if cell is empty and skip it if so
            if rowData.upper().startswith("TEMPLATE:"):
                if current_template_lines and current_template_name:
                    templates_dict[current_template_name] = "\n".join(
                        current_template_lines
                    )
                current_template_name = rowData.split(":")[1].strip()
                current_template_lines = []
            else:
                current_template_lines.append(rowData)
    # add last template
    if current_template_lines and current_template_name:
        templates_dict[current_template_name] = "\n".join(current_template_lines)

